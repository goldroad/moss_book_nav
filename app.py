import os
import sys
import sqlite3
from tempfile import NamedTemporaryFile
from flask import Flask, request, render_template, redirect, url_for, flash

# ========== PyInstaller 兼容 ==========
# 打包后资源释放到 sys._MEIPASS（临时目录），开发期用脚本所在目录
def _resolve_base_dir() -> str:
    """返回应用根目录（PyInstaller 兼容）"""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        # --onefile 模式：数据文件在 _MEIPASS，用户数据应复制到可写目录
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

BASE_APP_DIR = _resolve_base_dir()
# 当 frozen 时，用户可写数据（数据库、JSON）放在 exe 同目录
BASE_USER_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))

# 仅开发期切工作目录（打包后 Flask 会处理路径）
if not getattr(sys, "frozen", False):
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

# External deps (previously in main.py)
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from pypinyin import lazy_pinyin
except ImportError:
    lazy_pinyin = None

import json
import configparser
from typing import Optional, List, Tuple

# DB constants and schema (moved from main.py)
# frozen 模式下动态解析 DB 路径（因为 desktop_app.py 会在启动后 chdir）
def _resolve_db_path() -> str:
    """解析数据库路径，并在 frozen 模式下确保数据库存在"""
    if getattr(sys, "frozen", False):
        # 优先用环境变量
        env_db = os.environ.get("BOOKNAVI_DB")
        if env_db and os.path.exists(env_db):
            return env_db

        exe_dir_db = os.path.join(os.path.dirname(sys.executable), "booknavi.db")

        # 如果 exe 旁没有数据库或数据库过小(不含 moss_data 表)，从 _MEIPASS 复制
        need_copy = False
        if not os.path.exists(exe_dir_db):
            need_copy = True
        else:
            try:
                import sqlite3 as _sqlite3
                _conn = _sqlite3.connect(exe_dir_db)
                _tables = [r[0] for r in _conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
                _conn.close()
                if "moss_data" not in _tables:
                    need_copy = True
            except Exception:
                need_copy = True

        if need_copy and hasattr(sys, "_MEIPASS"):
            try:
                import shutil as _shutil
                bundled = os.path.join(sys._MEIPASS, "booknavi.db")
                if os.path.exists(bundled):
                    _shutil.copy2(bundled, exe_dir_db)
            except Exception:
                pass

        return exe_dir_db
    return os.path.join(BASE_USER_DIR, "booknavi.db")

DB_DEFAULT = _resolve_db_path()

SCHEMA_SQL = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS books (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    book_name TEXT NOT NULL,
    short_name TEXT NOT NULL UNIQUE,
    offset INTEGER DEFAULT 0,
    has_pdf INTEGER DEFAULT 1,
    region TEXT,
    publisher TEXT,
    publish_date TEXT,
    price REAL,
    page INTEGER,
    moss_count INTEGER DEFAULT 0,
    author TEXT,
    cover_image TEXT,
    notes TEXT,
    has_txt INTEGER DEFAULT 0,
    has_line INTEGER DEFAULT 0,
    has_spec INTEGER DEFAULT 0,
    has_env INTEGER DEFAULT 0,
    has_micro INTEGER DEFAULT 0,
    has_micro_section INTEGER DEFAULT 0,
    has_elec INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS mosses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    moss_name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS moss_pages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    book_id INTEGER NOT NULL,
    moss_id INTEGER NOT NULL,
    page_id TEXT NOT NULL,
    UNIQUE(book_id, moss_id, page_id),
    FOREIGN KEY(book_id) REFERENCES books(id) ON DELETE CASCADE,
    FOREIGN KEY(moss_id) REFERENCES mosses(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_moss_name ON mosses(moss_name);
CREATE INDEX IF NOT EXISTS idx_moss_pages_moss ON moss_pages(moss_id);
CREATE INDEX IF NOT EXISTS idx_moss_pages_book ON moss_pages(book_id);
"""

# DB helpers

def connect_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db(conn: sqlite3.Connection, clear: bool = False) -> None:
    conn.executescript(SCHEMA_SQL)
    # 确保 books 表存在 offset 列（默认 0）
    try:
        conn.execute("ALTER TABLE books ADD COLUMN offset INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    # 确保 books 表存在 has_pdf 列（1=有PDF，0=无PDF；默认有PDF）
    try:
        conn.execute("ALTER TABLE books ADD COLUMN has_pdf INTEGER DEFAULT 1")
    except sqlite3.OperationalError:
        pass
    # 新增书籍元数据列（若不存在则添加）
    for sql in [
        "ALTER TABLE books ADD COLUMN region TEXT",
        "ALTER TABLE books ADD COLUMN publisher TEXT",
        "ALTER TABLE books ADD COLUMN publish_date TEXT",
        "ALTER TABLE books ADD COLUMN price REAL",
        "ALTER TABLE books ADD COLUMN page INTEGER",
        "ALTER TABLE books ADD COLUMN moss_count INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN author TEXT",
        "ALTER TABLE books ADD COLUMN cover_image TEXT",
        "ALTER TABLE books ADD COLUMN notes TEXT",
        "ALTER TABLE books ADD COLUMN has_txt INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_line INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_spec INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_env INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_micro INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_micro_section INTEGER DEFAULT 0",
        "ALTER TABLE books ADD COLUMN has_elec INTEGER DEFAULT 0"
    ]:
        try:
            conn.execute(sql)
        except sqlite3.OperationalError:
            pass

    # 修复历史错误：moss_pages 的外键可能错误指向 books_old
    try:
        fk_rows = conn.execute("PRAGMA foreign_key_list(moss_pages)").fetchall()
        # row columns: (id, seq, table, from, to, on_update, on_delete, match)
        needs_fix = any(((r[2] if not isinstance(r, sqlite3.Row) else r["table"]) == "books_old") for r in fk_rows)
        if needs_fix:
            conn.execute("PRAGMA foreign_keys=OFF")
            conn.execute("BEGIN")
            conn.execute(
                """
                CREATE TABLE moss_pages_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    moss_id INTEGER NOT NULL,
                    page_id TEXT NOT NULL,
                    UNIQUE(book_id, moss_id, page_id),
                    FOREIGN KEY(book_id) REFERENCES books(id) ON DELETE CASCADE,
                    FOREIGN KEY(moss_id) REFERENCES mosses(id) ON DELETE CASCADE
                )
                """
            )
            conn.execute("INSERT INTO moss_pages_new(id, book_id, moss_id, page_id) SELECT id, book_id, moss_id, page_id FROM moss_pages")
            conn.execute("DROP TABLE moss_pages")
            conn.execute("ALTER TABLE moss_pages_new RENAME TO moss_pages")
            conn.execute("COMMIT")
            conn.execute("PRAGMA foreign_keys=ON")
            # 重新创建索引
            conn.execute("CREATE INDEX IF NOT EXISTS idx_moss_pages_moss ON moss_pages(moss_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_moss_pages_book ON moss_pages(book_id)")
    except sqlite3.OperationalError:
        # 若 moss_pages 不存在或 PRAGMA 失败，忽略
        pass

    if clear:
        # 清空页码与苔藓名称数据，但保留现有 books 映射
        conn.execute("DELETE FROM moss_pages")
        conn.execute("DELETE FROM mosses")
        # 重置 AUTOINCREMENT 计数，使下次从 1 开始
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('moss_pages','mosses')")
        except sqlite3.OperationalError:
            # sqlite_sequence 可能不存在或无 AUTOINCREMENT，忽略
            pass
        conn.commit()


def get_or_create_book(conn: sqlite3.Connection, short_name: str, book_name: str, offset: int = 0, has_pdf: int = 1) -> int:
    short_name = short_name.strip()
    book_name = book_name.strip()
    cur = conn.execute("SELECT id FROM books WHERE short_name = ?", (short_name,))
    row = cur.fetchone()
    if row:
        conn.execute("UPDATE books SET book_name = ?, offset = ?, has_pdf = ? WHERE id = ?", (book_name, offset, has_pdf, row["id"]))
        return row["id"]
    conn.execute("INSERT INTO books(book_name, short_name, offset, has_pdf) VALUES (?, ?, ?, ?)", (book_name, short_name, offset, has_pdf))
    return conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]


def get_or_create_moss(conn: sqlite3.Connection, moss_name: str) -> int:
    moss_name = moss_name.strip()
    cur = conn.execute("SELECT id FROM mosses WHERE moss_name = ?", (moss_name,))
    row = cur.fetchone()
    if row:
        return row["id"]
    conn.execute("INSERT OR IGNORE INTO mosses(moss_name) VALUES (?)", (moss_name,))
    return conn.execute("SELECT id FROM mosses WHERE moss_name = ?", (moss_name,)).fetchone()["id"]


def insert_page(conn: sqlite3.Connection, book_id: int, moss_id: int, page_id: str) -> None:
    page_id = page_id.strip()
    if not page_id:
        return
    conn.execute(
        "INSERT OR IGNORE INTO moss_pages(book_id, moss_id, page_id) VALUES (?, ?, ?)",
        (book_id, moss_id, page_id),
    )

# === 书籍导入（从 xlsx 导入到 books 表） ===
def import_books_xlsx(conn: sqlite3.Connection, xlsx_path: str, clear: bool = True) -> int:
    if load_workbook is None:
        raise RuntimeError("openpyxl 未安装，请先执行: pip install -r requirements.txt")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"未找到 Excel 文件: {xlsx_path}")

    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheet = wb.worksheets[0] if wb.worksheets else None
    if sheet is None:
        raise ValueError("Excel 内容为空")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 内容为空")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [h for h in headers if h]  # 去除空表头
    if not headers:
        raise ValueError("Excel 首行表头为空")

    table_cols = [r[1] for r in conn.execute("PRAGMA table_info(books)").fetchall()]
    if not table_cols:
        raise RuntimeError("数据库中不存在 books 表")

    unknown = [h for h in headers if h not in table_cols]
    if unknown:
        raise ValueError(f"Excel 表头包含未知列: {unknown}\n当前 books 表列: {table_cols}")

    if clear:
        # 清空旧数据（将级联删除 moss_pages 中关联记录）
        conn.execute("DELETE FROM books")
        # 重置自增序列，使新插入 ID 从 1 开始
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name='books'")
        except sqlite3.OperationalError:
            pass

    cols_sql = ", ".join(headers)
    placeholders = ", ".join(["?"] * len(headers))
    insert_sql = f"INSERT INTO books ({cols_sql}) VALUES ({placeholders})"

    to_insert = []
    for idx, row in enumerate(rows[1:], start=2):  # 从第2行开始
        # 截断到表头长度，过长的列忽略，不足的补 None
        row_values = (row or ())
        values = list(row_values[:len(headers)]) + [None] * (len(headers) - len(row_values))
        values = [v.strip() if isinstance(v, str) else v for v in values]
        to_insert.append(values)

    if to_insert:
        conn.executemany(insert_sql, to_insert)

    conn.commit()
    return len(to_insert)


def load_mapping(mapping_path: Optional[str]) -> dict:
    # 返回结构：{短名: {"book_name": ..., "offset": int, "has_pdf": int}}
    default_ini = os.path.join(os.path.dirname(__file__), "config.ini")

    def parse_offset(s: str) -> int:
        s = (s or "").strip()
        if not s:
            return 0
        try:
            return int(s)
        except ValueError:
            try:
                return int(s.replace("+", ""))
            except Exception:
                return 0

    def load_ini(path: str) -> dict:
        result = {}
        # 先尝试按标准 INI 解析；若失败则回退到逐行 CSV 解析
        try:
            cp = configparser.ConfigParser()
            with open(path, "r", encoding="utf-8") as f:
                cp.read_file(f)  # 可能抛出 ParsingError
            if cp.has_section("books"):
                # 支持 value 为 "书名" 或 "书名,偏移"
                for k, v in cp.items("books"):
                    short = k.strip()
                    parts = [p.strip() for p in v.split(",")]
                    if len(parts) >= 2:
                        book_name, off_str = parts[0], parts[1]
                        result[short] = {"book_name": book_name, "offset": parse_offset(off_str), "has_pdf": 1}
                    elif len(parts) == 1:
                        result[short] = {"book_name": parts[0], "offset": 0, "has_pdf": 0}
                return result
        except configparser.Error:
            pass

        # 回退：兼容纯文本行，每行格式为 "短名,书名,偏移" 或 "短名,书名"
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or line.startswith("["):
                    # 跳过空行、注释行、段落头（如 [books]）
                    continue
                parts = [p.strip() for p in line.split(",")]
                if len(parts) >= 3:
                    short, book_name, off_str = parts[0], parts[1], parts[2]
                    result[short] = {"book_name": book_name, "offset": parse_offset(off_str), "has_pdf": 1}
                elif len(parts) == 2:
                    short, book_name = parts[0], parts[1]
                    result[short] = {"book_name": book_name, "offset": 0, "has_pdf": 0}
        return result

    if not mapping_path:
        if os.path.exists(default_ini):
            return load_ini(default_ini)
        return {}

    if mapping_path.lower().endswith(".ini"):
        if not os.path.exists(mapping_path):
            raise FileNotFoundError(f"book mapping ini not found: {mapping_path}")
        return load_ini(mapping_path)

    if not os.path.exists(mapping_path):
        raise FileNotFoundError(f"book mapping file not found: {mapping_path}")
    with open(mapping_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        result = {}
        if isinstance(data, dict):
            for k, v in data.items():
                short = str(k).strip()
                if isinstance(v, dict):
                    book_name = str(v.get("book_name", short)).strip()
                    offset = parse_offset(str(v.get("offset", "0")))
                    has_pdf = 1 if int(v.get("has_pdf", 1)) else 0
                else:
                    book_name = str(v).strip()
                    offset = 0
                    has_pdf = 1
                result[short] = {"book_name": book_name, "offset": offset, "has_pdf": has_pdf}
            return result
        raise ValueError("book mapping must be a JSON object {short_name: book_name | {book_name, offset, has_pdf}}")


def load_mapping_from_db(conn: sqlite3.Connection) -> dict:
    rows = conn.execute("SELECT short_name, book_name, offset, has_pdf FROM books").fetchall()
    result = {}
    for r in rows:
        short = (r["short_name"] or "").strip()
        if not short:
            continue
        result[short] = {
            "book_name": (r["book_name"] or short).strip(),
            "offset": int(r["offset"]) if r["offset"] is not None else 0,
            "has_pdf": (int(r["has_pdf"]) if r["has_pdf"] is not None else 1),
        }
    return result


def _normalize_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.strip()


def import_xlsx(conn: sqlite3.Connection, xlsx_path: str, mapping_path: Optional[str]) -> Tuple[int, int, int]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Please install dependencies: pip install -r requirements.txt")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"xlsx file not found: {xlsx_path}")

    mapping = load_mapping_from_db(conn) if mapping_path is None else load_mapping(mapping_path)

    wb = load_workbook(filename=xlsx_path, data_only=True)
    total_books = 0
    total_mosses_inserted = 0
    total_pages_inserted = 0

    for sheet in wb.worksheets:
        short_name = _normalize_cell(sheet.title)
        info = mapping.get(short_name, {"book_name": short_name, "offset": 0, "has_pdf": 1})
        book_name = info["book_name"]
        offset = info.get("offset", 0)
        has_pdf = 1 if int(info.get("has_pdf", 1)) else 0
        book_id = get_or_create_book(conn, short_name, book_name, offset, has_pdf)
        total_books += 1

        for row in sheet.iter_rows(values_only=True):
            moss_name = _normalize_cell(row[0] if len(row) >= 1 else None)
            page = _normalize_cell(row[1] if len(row) >= 2 else None)

            if not moss_name and not page:
                continue

            if moss_name:
                moss_id = get_or_create_moss(conn, moss_name)
                total_mosses_inserted += 1
                if page:
                    insert_page(conn, book_id, moss_id, page)
                    total_pages_inserted += 1

    conn.commit()
    return total_books, total_mosses_inserted, total_pages_inserted


def to_pinyin(s: str) -> str:
    if not s:
        return ""
    if lazy_pinyin is None:
        return s  # fallback: original string order
    try:
        return "".join(lazy_pinyin(s))
    except Exception:
        return s


def search(conn: sqlite3.Connection, keyword: str, limit: Optional[int] = None, order_by_pinyin: bool = True) -> List[dict]:
    kw = f"%{keyword.strip()}%"
    moss_rows = conn.execute(
        "SELECT id, moss_name FROM mosses WHERE moss_name LIKE ?",
        (kw,),
    ).fetchall()

    moss_items = [(row["id"], row["moss_name"]) for row in moss_rows]
    if order_by_pinyin:
        moss_items.sort(key=lambda x: to_pinyin(x[1]))

    def compute_pdf_page(page: str, offset: int):
        import re
        s = (page or "").strip()
        m = re.match(r"^(\d+)", s)
        if not m:
            return None
        try:
            return str(int(m.group(1)) + int(offset))
        except Exception:
            return None

    results = []
    count = 0
    for moss_id, moss_name in moss_items:
        pages = conn.execute(
            """
            SELECT b.id AS book_id, b.short_name, b.book_name, b.offset AS offset, b.has_pdf AS has_pdf, mp.page_id
            FROM moss_pages mp
            JOIN books b ON b.id = mp.book_id
            WHERE mp.moss_id = ?
            ORDER BY b.short_name COLLATE NOCASE, mp.page_id COLLATE NOCASE
            """,
            (moss_id,),
        ).fetchall()
        if not pages:
            continue
        entry = {
            "moss_name": moss_name,
            "locations": [
                {
                    "book_id": p["book_id"],
                    "short_name": p["short_name"],
                    "book_name": p["book_name"],
                    "page": p["mp.page_id"] if "mp.page_id" in p.keys() else p["page_id"],
                    "offset": p["offset"],
                    "has_pdf": p["has_pdf"],
                    "pdf_page": (compute_pdf_page(p["mp.page_id"] if "mp.page_id" in p.keys() else p["page_id"], p["offset"]) if int(p["has_pdf"]) and int(p["offset"] or 0) != 0 else None),
                }
                for p in pages
            ],
        }
        results.append(entry)
        count += 1
        if limit and count >= limit:
            break
    return results

app = Flask(__name__,
            template_folder=os.path.join(BASE_APP_DIR, "templates"),
            static_folder=os.path.join(BASE_APP_DIR, "static"),
            static_url_path="/static")
app.secret_key = os.environ.get("BOOKNAVI_SECRET", "booknavi-secret")

DB_PATH = os.environ.get("BOOKNAVI_DB", DB_DEFAULT)

# 记录与使用 PID 文件（放在用户可写目录）
PID_FILE = os.path.join(
    os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_USER_DIR,
    "app.pid"
)

# Ensure DB schema exists at startup
with connect_db(DB_PATH) as conn:
    init_db(conn, clear=False)

# === 名录脑图版：启动时预加载 moss2026.json ===
MOSS_JSON_PATH = os.path.join(BASE_APP_DIR, "moss2026.json")

def _prune_null_leaves(node):
    if isinstance(node, dict):
        pruned = {}
        for k, v in node.items():
            if isinstance(v, dict):
                pruned[k] = _prune_null_leaves(v)
            # 值为 null（叶子）不展示
        return pruned
    return {}

def _build_name_index(raw_node, path, index):
    if isinstance(raw_node, dict):
        for k, v in raw_node.items():
            if isinstance(v, dict):
                index.setdefault(k, []).append({"path": path + [k], "leaf": False})
                _build_name_index(v, path + [k], index)
            else:
                # 叶子：值为 null
                index.setdefault(k, []).append({"path": path + [k], "leaf": True})

try:
    with open(MOSS_JSON_PATH, "r", encoding="utf-8") as f:
        _raw_mindmap = json.load(f)
except Exception:
    _raw_mindmap = {}

MINDMAP_TREE = _raw_mindmap if isinstance(_raw_mindmap, dict) else {}
MINDMAP_INDEX = {}
if isinstance(_raw_mindmap, dict):
    _build_name_index(_raw_mindmap, [], MINDMAP_INDEX)

@app.route("/mindmap")
def mindmap():
    return render_template(
        "mindmap.html",
        mindmap_data=json.dumps(MINDMAP_TREE, ensure_ascii=False),
        mindmap_index=json.dumps(MINDMAP_INDEX, ensure_ascii=False),
    )

@app.route("/")
def index():
    keyword = request.args.get("q", "").strip()
    limit_raw = request.args.get("limit", "").strip()

    # 多选书籍过滤（默认全选）
    books = list_books()
    selected_ids = request.args.getlist("books")
    if not selected_ids:
        selected_ids = [str(b["id"]) for b in books]

    limit = None
    if limit_raw:
        try:
            limit = int(limit_raw)
        except ValueError:
            flash("limit 参数必须为整数", "error")

    results = []
    groups = []
    moss_total = None
    if keyword:
        with connect_db(DB_PATH) as conn:
            init_db(conn, clear=False)
            results = search(conn, keyword, limit=limit, order_by_pinyin=True)
        # 先按所选书籍过滤位置；若某苔藓在所选书籍中无位置则移除
        filtered = []
        for item in results:
            locs = [loc for loc in item["locations"] if str(loc.get("book_id")) in selected_ids]
            if locs:
                filtered.append({"moss_name": item["moss_name"], "locations": locs})
        # 按书籍聚合（每个书籍下展示苔藓名称及其页码列表）
        grouped = {}
        for item in filtered:
            moss_name = item["moss_name"]
            for loc in item["locations"]:
                bid = loc["book_id"]
                bname = loc["book_name"]
                offset = int(loc.get("offset", 0))
                has_pdf = int(loc.get("has_pdf", 1))
                page = (loc["page"] or "").strip()
                g = grouped.setdefault(bid, {"book_id": bid, "book_name": bname, "offset": offset, "has_pdf": has_pdf, "mosses": {}})
                moss_pages = g["mosses"].setdefault(moss_name, set())
                if page:
                    moss_pages.add(page)
        groups = []
        for g in grouped.values():
            offset = int(g.get("offset", 0))
            has_pdf = int(g.get("has_pdf", 1))
            moss_entries = [{
                "name": name,
                "page_items": [{"page": p, "pdf": (str(int(p) + offset) if has_pdf and p.isdigit() else (p if has_pdf and offset == 0 else None))} for p in sorted(pages, key=page_sort_key)],
            } for name, pages in g["mosses"].items()]
            moss_entries.sort(key=lambda e: page_sort_key(e["page_items"][0]["page"]) if e["page_items"] else (10**9, ""))
            groups.append({
                "book_id": g["book_id"],
                "book_name": g["book_name"],
                "has_pdf": has_pdf,
                "mosses": moss_entries,
                "count": len(moss_entries),
            })
        # 书籍组按全称排序
        groups.sort(key=lambda x: x["book_name"])
        # 补充封面图到分组
        cover_map = {b["id"]: b.get("cover_image") for b in books}
        for g in groups:
            g["cover_image"] = cover_map.get(g["book_id"])  # 供模板显示缩略图
        moss_total = sum(g["count"] for g in groups)
        results = filtered  # 保留以兼容旧模板变量（现模板使用 groups）
    return render_template(
        "index.html",
        keyword=keyword,
        results=results,
        limit=limit,
        books=books,
        selected_ids=selected_ids,
        groups=groups,
        moss_total=moss_total,
    )


@app.route("/importpages", methods=["GET", "POST"])
def import_pages():
    if request.method == "GET":
        return render_template("importpages.html")

    clear = request.form.get("clear") == "on"
    xlsx_file = request.files.get("xlsx_file")

    if not xlsx_file or not xlsx_file.filename.lower().endswith(".xlsx"):
        flash("请上传 .xlsx 文件", "error")
        return redirect(url_for("import_pages"))

    temp_xlsx = NamedTemporaryFile(delete=False, suffix=".xlsx")
    xlsx_path = temp_xlsx.name
    xlsx_file.save(xlsx_path)

    mapping_path = None  # 使用数据库 books 表中的短名→书名映射

    try:
        with connect_db(DB_PATH) as conn:
            init_db(conn, clear=clear)
            books, mosses, pages = import_xlsx(conn, xlsx_path, mapping_path)
        flash(f"导入完成：书籍={books}，苔藓记录(含去重插入)={mosses}，页码记录={pages}。", "success")
    except Exception as e:
        flash(f"导入失败：{e}", "error")
    finally:
        try:
            os.unlink(xlsx_path)
        except Exception:
            pass

    return redirect(url_for("import_pages"))

def list_books():
    with connect_db(DB_PATH) as conn:
        init_db(conn, clear=False)
        rows = conn.execute(
            """
            SELECT b.id, b.short_name, b.book_name, b.has_pdf, b.offset, b.cover_image, COALESCE(cnt.c, 0) AS data_count
            FROM books b
            LEFT JOIN (
                SELECT book_id, COUNT(*) AS c
                FROM moss_pages
                GROUP BY book_id
            ) cnt ON cnt.book_id = b.id
            ORDER BY b.short_name COLLATE NOCASE
            """
        ).fetchall()
        return [
            {
                "id": r["id"],
                "short_name": r["short_name"],
                "book_name": r["book_name"],
                "has_pdf": (int(r["has_pdf"]) if r["has_pdf"] is not None else 0),
                "offset": (int(r["offset"]) if r["offset"] is not None else 0),
                "cover_image": r["cover_image"],
                "data_count": (int(r["data_count"]) if r["data_count"] is not None else 0)
            }
            for r in rows
        ]


def page_sort_key(s: str):
    import re
    s = (s or "").strip()
    m = re.match(r"^(\d+)", s)
    if m:
        num = int(m.group(1))
        rest = s[m.end():].strip().lower()
    else:
        num = 10**9  # 非数字页码排在最后
        rest = s.lower()
    return (num, rest)


@app.route("/browse")
def browse():
    books = list_books()
    sel_id_raw = request.args.get("book_id", "").strip()
    selected_id = None
    selected_book = None
    if sel_id_raw:
        for b in books:
            if str(b["id"]) == sel_id_raw:
                selected_id = b["id"]
                selected_book = b
                break
    else:
        if books:
            selected_book = min(books, key=lambda x: x["id"])  # 默认选择ID最小的书籍
            selected_id = selected_book["id"]

    mosses = []
    if selected_book:
        with connect_db(DB_PATH) as conn:
            init_db(conn, clear=False)
            rows = conn.execute(
                """
                SELECT mp.page_id AS page, m.moss_name AS moss, b.offset AS offset, b.has_pdf AS has_pdf
                FROM moss_pages mp
                JOIN mosses m ON m.id = mp.moss_id
                JOIN books b ON b.id = mp.book_id
                WHERE mp.book_id = ?
                """,
                (selected_book["id"],),
            ).fetchall()
        by_moss = {}
        book_offset = 0
        book_has_pdf = 1
        for r in rows:
            p = (r["page"] or "").strip()
            moss = r["moss"]
            book_offset = int(r["offset"]) if r["offset"] is not None else 0
            book_has_pdf = int(r["has_pdf"]) if r["has_pdf"] is not None else 1
            if p:
                by_moss.setdefault(moss, set()).add(p)
        for name, pages in by_moss.items():
            items = []
            for p in sorted(pages, key=page_sort_key):
                pdf = None
                if book_has_pdf and book_offset != 0 and p.isdigit():
                    try:
                        pdf = str(int(p) + book_offset)
                    except Exception:
                        pdf = None
                items.append({"page": p, "pdf": pdf})
            mosses.append({"name": name, "page_items": items})
        mosses.sort(key=lambda e: page_sort_key(e["page_items"][0]["page"]) if e["page_items"] else (10**9, ""))

    return render_template(
        "browse.html",
        books=books,
        selected_id=selected_id,
        selected_book=selected_book,
        mosses=mosses,
    )


# === 新增：书籍目录与封面图片静态路由 ===
from flask import send_from_directory
IMG_DIR = os.path.join(BASE_APP_DIR, "img")

@app.route("/img/<path:filename>")
def img_file(filename):
    return send_from_directory(IMG_DIR, filename)

@app.route("/delete_book/<int:book_id>", methods=["POST"])
def delete_book(book_id):
    conn = connect_db(DB_DEFAULT)
    try:
        # 删除 moss_pages 关联记录
        conn.execute("DELETE FROM moss_pages WHERE book_id = ?", (book_id,))
        # 删除 books 记录
        conn.execute("DELETE FROM books WHERE id = ?", (book_id,))
        conn.commit()
        flash("删除成功。", "success")
    except Exception as e:
        flash(f"删除失败：{e}", "error")
    finally:
        conn.close()
    return redirect(url_for("catalog"))

@app.route("/catalog")
def catalog():
    # Filter parameters
    sel_regions = request.args.getlist("region")
    sel_has_txt = request.args.get("has_txt", "")
    sel_has_line = request.args.get("has_line", "")
    sel_has_spec = request.args.get("has_spec", "")
    sel_has_env = request.args.get("has_env", "")
    sel_has_micro = request.args.get("has_micro", "")
    sel_has_micro_section = request.args.get("has_micro_section", "")
    sel_has_elec = request.args.get("has_elec", "")
    sort_by = request.args.get("sort_by", "date_desc")  # default, date_asc, date_desc

    with connect_db(DB_PATH) as conn:
        init_db(conn, clear=False)

        # Get all regions for the filter UI
        all_regions = [r["region"] for r in conn.execute("SELECT DISTINCT region FROM books WHERE region IS NOT NULL AND region != '' ORDER BY region").fetchall()]

        # Build query
        sql = """
            SELECT b.id, b.short_name, b.book_name, b.has_pdf, b.region, b.publisher, b.publish_date, b.price, b.page, b.moss_count, b.author, b.cover_image, b.notes,
                   b.has_txt, b.has_line, b.has_spec, b.has_env, b.has_micro, b.has_micro_section, b.has_elec,
                   COALESCE(cnt.c, 0) AS data_count
            FROM books b
            LEFT JOIN (
                SELECT book_id, COUNT(*) AS c
                FROM moss_pages
                GROUP BY book_id
            ) cnt ON cnt.book_id = b.id
            WHERE 1=1
        """
        params = []

        if sel_regions:
            placeholders = ",".join(["?"] * len(sel_regions))
            sql += f" AND b.region IN ({placeholders})"
            params.extend(sel_regions)

        # Helper for tri-state filters
        def add_filter(field, val):
            nonlocal sql
            if val and val.isdigit():
                sql += f" AND b.{field} = ?"
                params.append(int(val))

        add_filter("has_txt", sel_has_txt)
        add_filter("has_line", sel_has_line)
        add_filter("has_spec", sel_has_spec)
        add_filter("has_env", sel_has_env)
        add_filter("has_micro", sel_has_micro)
        add_filter("has_micro_section", sel_has_micro_section)
        add_filter("has_elec", sel_has_elec)

        # Execute query
        rows = conn.execute(sql, params).fetchall()
        
        books = []
        for r in rows:
            books.append({
                "id": r["id"],
                "book_name": r["book_name"],
                "has_pdf": int(r["has_pdf"]) if r["has_pdf"] is not None else 0,
                "region": r["region"],
                "publisher": r["publisher"],
                "publish_date": r["publish_date"],
                "price": r["price"],
                "page": r["page"],
                "moss_count": r["moss_count"],
                "has_txt": int(r["has_txt"]) if r["has_txt"] is not None else 0,
                "has_line": int(r["has_line"]) if r["has_line"] is not None else 0,
                "has_spec": int(r["has_spec"]) if r["has_spec"] is not None else 0,
                "has_env": int(r["has_env"]) if r["has_env"] is not None else 0,
                "has_micro": int(r["has_micro"]) if r["has_micro"] is not None else 0,
                "has_micro_section": int(r["has_micro_section"]) if r["has_micro_section"] is not None else 0,
                "has_elec": int(r["has_elec"]) if r["has_elec"] is not None else 0,
                "author": r["author"],
                "cover_image": r["cover_image"],
                "notes": r["notes"],
                "data_count": (int(r["data_count"]) if r["data_count"] is not None else 0),
                "short_name": r["short_name"]
            })

        # Python-side sorting
        if sort_by == "date_asc":
            # Extract year/number from publish_date for sorting
            # Handle cases like "1994", "2006.5"
            def date_key(b):
                d = str(b["publish_date"] or "")
                import re
                m = re.search(r"(\d+(\.\d+)?)", d)
                if m:
                    return float(m.group(1))
                return 99999.0 # Unknown dates last
            books.sort(key=date_key)
        elif sort_by == "date_desc":
            def date_key(b):
                d = str(b["publish_date"] or "")
                import re
                m = re.search(r"(\d+(\.\d+)?)", d)
                if m:
                    return float(m.group(1))
                return -1.0 # Unknown dates last (or first? usually last is better)
            books.sort(key=date_key, reverse=True)
        else:
            # Default sort by short_name
            books.sort(key=lambda x: (x["short_name"] or "").lower())

    return render_template(
        "catalog.html", 
        books=books, 
        all_regions=all_regions,
        sel_regions=sel_regions,
        sel_has_txt=sel_has_txt,
        sel_has_line=sel_has_line,
        sel_has_spec=sel_has_spec,
        sel_has_env=sel_has_env,
        sel_has_micro=sel_has_micro,
        sel_has_micro_section=sel_has_micro_section,
        sel_has_elec=sel_has_elec,
        sort_by=sort_by
    )

@app.route("/importbooks")
def importbooks():
    return render_template("importbooks.html")

@app.route("/catalog/import", methods=["POST"])
def import_catalog_books():
    clear = request.form.get("clear") == "on"
    xlsx_file = request.files.get("books_xlsx")
    if not xlsx_file or not xlsx_file.filename.lower().endswith(".xlsx"):
        flash("请上传 .xlsx 文件", "error")
        return redirect(url_for("importbooks"))

    temp_xlsx = NamedTemporaryFile(delete=False, suffix=".xlsx")
    xlsx_path = temp_xlsx.name
    xlsx_file.save(xlsx_path)

    try:
        with connect_db(DB_PATH) as conn:
            init_db(conn, clear=False)
            count = import_books_xlsx(conn, xlsx_path, clear=clear)
        msg = f"导入书籍完成：{count} 条记录。"
        if clear:
            msg += " 已清空旧数据。"
        flash(msg, "success")
        return redirect(url_for("catalog"))
    except Exception as e:
        flash(f"导入书籍失败：{e}", "error")
        return redirect(url_for("importbooks"))
    finally:
        try:
            os.unlink(xlsx_path)
        except Exception:
            pass

# === Taxonomy database (moss_data) ===
# Chinese rank columns used for cascading selects
TAXO_CN_COLUMNS = [
    "门中文名",
    "纲中文名",
    "目中文名",
    "科中文名",
    "属中文名",
]

ALL_MOSS_COLUMNS = [
    "门拉丁名",
    "门中文名",
    "纲拉丁名",
    "纲中文名",
    "目拉丁名",
    "目中文名",
    "科拉丁名",
    "科中文名",
    "属拉丁名",
    "属中文名",
    "物种拉丁名",
    "物种中文名",
]


def distinct_values(conn: sqlite3.Connection, column: str, filters: dict) -> List[str]:
    where_parts = []
    params = []
    for k, v in filters.items():
        if v:
            where_parts.append(f'"{k}" = ?')
            params.append(v)
    sql = f'SELECT DISTINCT "{column}" AS val FROM moss_data'
    if where_parts:
        sql += ' WHERE ' + ' AND '.join(where_parts)
    sql += ' ORDER BY val COLLATE NOCASE'
    rows = conn.execute(sql, params).fetchall()
    return [r["val"] for r in rows if r["val"]]


def _build_taxonomy_conditions(selections: dict, keyword: str):
    where_parts = []
    params = []
    for col, val in selections.items():
        if val:
            where_parts.append(f'"{col}" = ?')
            params.append(val)
    if keyword:
        like = f"%{keyword}%"
        or_parts = [f'"{c}" LIKE ?' for c in ALL_MOSS_COLUMNS]
        where_parts.append('(' + ' OR '.join(or_parts) + ')')
        params.extend([like] * len(ALL_MOSS_COLUMNS))
    return where_parts, params


def count_taxonomy(conn: sqlite3.Connection, selections: dict, keyword: str) -> int:
    where_parts, params = _build_taxonomy_conditions(selections, keyword)
    sql = 'SELECT COUNT(*) AS c FROM moss_data'
    if where_parts:
        sql += ' WHERE ' + ' AND '.join(where_parts)
    row = conn.execute(sql, params).fetchone()
    return int(row["c"]) if row and row["c"] is not None else 0


def fetch_taxonomy(conn: sqlite3.Connection, selections: dict, keyword: str, limit: int, offset: int):
    where_parts, params = _build_taxonomy_conditions(selections, keyword)
    sql = 'SELECT ' + ', '.join([f'"{c}"' for c in ALL_MOSS_COLUMNS]) + ' FROM moss_data'
    if where_parts:
        sql += ' WHERE ' + ' AND '.join(where_parts)
    sql += ' ORDER BY "门拉丁名" COLLATE NOCASE, "纲拉丁名" COLLATE NOCASE, "目拉丁名" COLLATE NOCASE, "科拉丁名" COLLATE NOCASE, "属拉丁名" COLLATE NOCASE, "物种拉丁名" COLLATE NOCASE'
    sql += ' LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    return conn.execute(sql, params).fetchall()


def build_pagination(current_page: int, total_pages: int, window: int = 2):
    pages = []
    if total_pages <= 1:
        return [1]
    start = max(1, current_page - window)
    end = min(total_pages, current_page + window)
    if start > 1:
        pages.append(1)
        if start > 2:
            pages.append(None)  # 省略号
    for p in range(start, end + 1):
        pages.append(p)
    if end < total_pages:
        if end < total_pages - 1:
            pages.append(None)  # 省略号
        pages.append(total_pages)
    return pages


@app.route("/taxonomy", methods=["GET"])
def taxonomy():
    q = request.args.get("q", "").strip()
    gate = request.args.get("gate", "").strip()
    clazz = request.args.get("clazz", "").strip()
    order = request.args.get("order", "").strip()
    family = request.args.get("family", "").strip()
    genus = request.args.get("genus", "").strip()

    per_page_raw = request.args.get("per_page", "").strip()
    page_raw = request.args.get("page", "").strip()
    allowed_page_sizes = [100, 200, 500, 1000]
    per_page = 100
    try:
        if per_page_raw:
            v = int(per_page_raw)
            if v in allowed_page_sizes:
                per_page = v
    except ValueError:
        pass
    page = 1
    try:
        if page_raw:
            v = int(page_raw)
            if v >= 1:
                page = v
    except ValueError:
        pass

    with connect_db(DB_PATH) as conn:
        init_db(conn, clear=False)
        gates = distinct_values(conn, "门中文名", {})
        classes = distinct_values(conn, "纲中文名", ({"门中文名": gate} if gate else {}))
        order_filters = {k: v for k, v in {"门中文名": gate, "纲中文名": clazz}.items() if v}
        orders = distinct_values(conn, "目中文名", order_filters)
        family_filters = {k: v for k, v in {"门中文名": gate, "纲中文名": clazz, "目中文名": order}.items() if v}
        families = distinct_values(conn, "科中文名", family_filters)
        genus_filters = {k: v for k, v in {"门中文名": gate, "纲中文名": clazz, "目中文名": order, "科中文名": family}.items() if v}
        genera = distinct_values(conn, "属中文名", genus_filters)

        selections = {
            "门中文名": gate,
            "纲中文名": clazz,
            "目中文名": order,
            "科中文名": family,
            "属中文名": genus,
        }

        total_count = count_taxonomy(conn, selections, q)
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page
        rows = fetch_taxonomy(conn, selections, q, per_page, offset)
        pagination_pages = build_pagination(page, total_pages, window=2)

    return render_template(
        "taxonomy.html",
        q=q,
        gate=gate,
        clazz=clazz,
        order=order,
        family=family,
        genus=genus,
        gates=gates,
        classes=classes,
        orders=orders,
        families=families,
        genera=genera,
        rows=rows,
        columns=ALL_MOSS_COLUMNS,
        total_count=total_count,
        total_pages=total_pages,
        page=page,
        per_page=per_page,
        allowed_page_sizes=allowed_page_sizes,
        pagination_pages=pagination_pages,
    )

@app.route("/restart", methods=["POST"])
def restart():
    import threading, time, subprocess, signal
    pid = None
    try:
        with open(PID_FILE, "r", encoding="utf-8") as f:
            pid = int((f.read() or "").strip())
    except Exception:
        flash("未找到或无法读取 PID 文件，重启失败。", "error")
        return redirect(url_for("index"))

    def _delayed_kill(target_pid: int):
        time.sleep(0.5)
        try:
            if os.name == "nt":
                subprocess.run(["taskkill", "/PID", str(target_pid), "/F"], check=False)
            else:
                os.kill(target_pid, signal.SIGTERM)
        except Exception:
            pass

    threading.Thread(target=_delayed_kill, args=(pid,), daemon=True).start()
    flash("已发出重启指令，程序将很快关闭。请手动启动程序。", "success")
    return redirect(url_for("index"))

@app.route("/phylogeny")
def phylogeny():
    return render_template("phylogeny.html")

@app.route("/resources")
def resources():
    return render_template("resources.html")

@app.route("/about")
def about():
    return render_template("about.html")

def run_server(port: int = 5000):
    """启动 Flask 开发服务器（供桌面启动器调用）"""
    import socket

    def is_port_in_use(p: int) -> bool:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(1)
            try:
                return s.connect_ex(("127.0.0.1", p)) == 0
            except OSError:
                return False

    if is_port_in_use(port):
        print(f"[warn] 端口 {port} 已被占用，跳过启动")
        return

    try:
        with open(PID_FILE, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
    except Exception:
        pass

    print(f"[启动] 正在启动服务，端口: {port}...")
    app.run(host="127.0.0.1", port=port, debug=False)


if __name__ == "__main__":
    import sys as _sys
    # 打包后由 desktop_app.py 拉起；开发期直接运行 app.py 则打开浏览器
    if getattr(_sys, "frozen", False):
        # 打包后：仅启动 Flask，不打开浏览器
        run_server(5000)
    else:
        # 开发期：启动 Flask + 自动打开浏览器
        import webbrowser
        import threading

        def _open_browser():
            try:
                webbrowser.open("http://127.0.0.1:5000/")
            except Exception:
                pass

        threading.Timer(1.0, _open_browser).start()
        run_server(5000)