# 安装依赖（仅首次需要）：
# pip install openpyxl

import os
from openpyxl import load_workbook, Workbook

INPUT_FILE  = 'input.xlsx'
OUTPUT_FILE = 'output-with.xlsx'

def merge_adjacent_cols(src_file, dst_file):
    if not os.path.exists(src_file):
        raise FileNotFoundError(f'找不到文件：{src_file}')

    wb = load_workbook(src_file)
    new_wb = Workbook()
    # 删除默认空工作表
    new_wb.remove(new_wb.active)

    for sheet in wb:
        new_ws = new_wb.create_sheet(title=sheet.title)
        max_row = sheet.max_row
        # 一共 14 列，两两合并 → 7 列
        for col_idx in range(1, 14, 2):          # 1,3,5,7,9,11,13
            left_col  = col_idx
            right_col = col_idx + 1
            out_col   = (col_idx + 1) // 2       # 输出列号 1~7
            for r in range(1, max_row + 1):
                left_val  = sheet.cell(row=r, column=left_col).value  or  ''
                right_val = sheet.cell(row=r, column=right_col).value or  ''
                merged    = f"{left_val}_{right_val}"
                new_ws.cell(row=r, column=out_col, value=merged)

    new_wb.save(dst_file)
    print(f'合并完成，结果已保存到：{dst_file}')

if __name__ == '__main__':
    merge_adjacent_cols(INPUT_FILE, OUTPUT_FILE)